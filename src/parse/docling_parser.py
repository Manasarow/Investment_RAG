"""
Docling-based parser for SEC filings (HTM, PDF, DOCX, XLSX).
Extracts prose blocks, tables, and figure captions with page provenance.
Outputs a structured JSON file per document.

── Original bug-tracker fixes (#-prefix) ─────────────────────────────────────
  #10  OCR enabled for PDFs; HTM/DOCX skip OCR via per-format options
  #9   Manifest O(N²) eliminated: load once, mutate, write once per run
  #72  Tables encode period_type / period_signals by scanning header cells
  #11  Converter instantiated once per run (module-level singleton)
  #12  Table section tracked inline during doc.iterate_items()
  #13  Multi-row table headers merged into composite column names
  #73  statement_type classification (score-based, not first-match)
  #74  table_title extracted from nearest preceding SectionHeaderItem / caption
  #75  Structured table preserved: row_labels, column_headers, cell_grid, periods
  #14  Figure metadata recorded honestly (caption / alt-text / dims)
  #76  is_financial_statement / is_cover_page flags added to every table record
  #15  is_xbrl_noise improved
  #16  Minimum prose-block length lowered from 20 to 8 chars
  #17  Content sanity check warns on sparse PDF extraction
  #18  Excel (.xlsx) handled via openpyxl; .xls explicitly rejected

── Post-review round-1 corrections (R-prefix) ────────────────────────────────
  R1   Manifest write moved to single flush at end of run()
  R2   .xls explicitly rejected
  R3   Figure comment corrected to reflect actual implementation
  R4   Statement classification now score-based
  R5   Header/data boundary uses numeric-density
  R6   period_type extended to 'instant', 'mixed'
  R7   Row-label column inferred (leftmost mostly-text column)
  R8   cell_grid uses inferred label-column index
  R9   build_table_record passes pre-computed header/data_rows to renderer
  R10  is_cover_page kept as soft flag
  R11  item type checks use isinstance()
  R12  caption / text fields normalised through _safe_text()
  R13  page_start / page_end / pages provenance
  R14  save_parsed uses atomic .tmp -> replace()
  R15  raw_path resolved for absolute or relative manifest paths

── Sample-validated round-2 fixes (V-prefix) ─────────────────────────────────
  V1   XBRL filter: dual-guard catches namespace-dense inline XBRL blocks
  V2   SEC HTM headings (PART/Item/Note) promoted to section updates
  V3   page_is_surrogate flag for HTM files
  V4   Absorbed numeric header cells ejected to synthetic first data row
  V5   Standalone unit tokens ($, %) stripped from header, stored in units field
  V6   Footer/boilerplate prose blocks filtered

── Round-3 sample-validated fixes (W-prefix, confirmed by AAPL 8-K/10-Q) ────
  W1   Image filename prose blocks filtered (aapl-20230202_g1.jpg etc.)
  W2   Empty-caption figure placeholders filtered (no information content)
  W3   Filing-layout cover-page prose fragments filtered
        (e.g. "(State or other jurisdiction", "File Number)", standalone labels)
  W4   Financial-statement title TextItems detected and used as pending_table_title
        ("Apple Inc. CONDENSED CONSOLIDATED STATEMENTS OF...")
  W5   Comprehensive income added to statement_type scoring
  W6   _FIN_INTEGER_RE fixed: now matches comma-formatted numbers (72,994 etc.)
  W7   Phantom empty columns dropped after header cleanup
        (empty-header cols where all data values are also empty)
  W8   Standalone unit tokens ($, %) filtered from data cell values in cell_grid
  W9   Maturity/commitment schedules detected as period_type='maturity_schedule'
  W10  Securities-registration tables classified as cover_admin_table
  W11  Equity header_pat tightened to avoid matching securities tables
  W12  Checkbox / filer-status tables classified as cover_admin_table
  W13  TOC tables classified as cover_admin_table

── Round-5 sample-validated fixes (Y-prefix, confirmed by AAPL 8-K/10-Q output) ─
  Y1   Signature / filing dates excluded from period_cols: headers matching
        'Month D, YYYY' (no embedded newline) are not reporting-period columns.
        Genuine period column headers always carry the year on a second line.
  Y2   Column compaction: empty-header columns that carry data are folded into
        their preceding non-empty period-column. This fixes all cell_grid entries
        with col_header='' in core financial statements (income, balance sheet,
        equity, cash flows). Root cause: Docling rows had [$, value, $, value]
        pattern; V5 stripped $ from headers leaving '' alongside real values.
  Y3   is_cover_page forced True for cover_admin_table: all cover/admin tables
        are definitionally cover-page content regardless of numeric content.
  Y4   Column-header word-fragment spill filter extended: 'Adjusted', 'Unrealized',
        'Marketable', 'Non-Current', 'Equivalents', 'Fair Value', 'Cash and',
        'Gains', 'Losses', 'Cost' now suppressed. Verified zero false-positive risk.
  Y5   Note heading over-capture fixed: _extract_short_heading now calls
        _extract_note_title() for Note headings, stopping at the first
        body-text trigger word (The, This, During, In, As, etc.) using a
        space-preceded word boundary to avoid cutting mid-word (e.g. 'Instruments').
        'Note 9 - Contingencies The Company...' -> 'Note 9 - Contingencies'.
  X1   Section reset on financial-statement title: when _extract_fin_stmt_title()
        fires, current_section is set to "Item 1 - Financial Statements",
        fixing the TOC 'Item 1A.' bleed that tagged all primary statements wrong.
  X2   Table-spill prose filtered: parenthesized negative numbers ( 1,766 ) and
        bare date fragments (December 31,) are now suppressed before indexing.
        Confirmed zero false-positive risk on full 10-Q prose corpus.
  X3   Cover fragment pattern broadened: 'of incorporation or organization)'
        now filtered (was missing the 'or organization' variant).
  X4   cover_admin_table periods cleared: tables classified as cover_admin_table
        always get periods=[] to prevent year-like strings in debt maturity names
        (e.g. "Notes due 2024") from being misread as period columns.
  X5   Signature/date table periods cleared: _PERIOD_INSTANT_RE tightened to
        require 'as of' or 'balance at' context; bare 'Date: February 2, 2023'
        headers no longer produce spurious period entries.
  X6   Hierarchy distortion fixed: _render_table_text now receives label_col and
        skips using the label column's header as a row-value prefix. Previously
        'ASSETS: / Current assets: / Cash and cash equivalents' was prepended to
        every balance sheet data row, even rows in liabilities or equity sections.

Known limitations (deferred to later pipeline stages):
  - Unit / scale / currency normalisation belongs in the chunker.
  - Table continuation across pages requires Docling cross-item state.
  - Row hierarchy (parent/subtotal/child rows) belongs in the chunker.
  - HTM files have no page concept; block_order is the position signal.
  - True visual understanding of charts (VLM/OCR) not implemented.
  - Column-header word fragments ('Weighted-Average', 'Grant Date Fair' etc.)
    require table-adjacency context to suppress safely; deferred to chunker.
  - Securities registration row_labels all '-': multi-row header merge absorbs
    the trading symbol column; cover_admin_table tables are not retrieval targets.
"""

import re
import json
import logging
import threading
from pathlib import Path
from typing import Optional

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

BASE_DIR   = Path(__file__).resolve().parents[2]
PARSED_DIR = BASE_DIR / "data" / "parsed"
TABLES_DIR = BASE_DIR / "data" / "tables"
MANIFEST   = BASE_DIR / "data_manifest" / "manifest.jsonl"

PARSED_DIR.mkdir(parents=True, exist_ok=True)
TABLES_DIR.mkdir(parents=True, exist_ok=True)

_MANIFEST_LOCK = threading.Lock()

# ── Lazy Docling type imports ──────────────────────────────────────────────────
_DOCLING_TYPES: dict = {}

def _get_docling_types() -> dict:
    global _DOCLING_TYPES
    if _DOCLING_TYPES:
        return _DOCLING_TYPES
    try:
        from docling.datamodel.document import (
            SectionHeaderItem, TextItem, TableItem, PictureItem,
        )
        _DOCLING_TYPES = {
            "SectionHeaderItem": SectionHeaderItem,
            "TextItem":          TextItem,
            "TableItem":         TableItem,
            "PictureItem":       PictureItem,
        }
    except ImportError:
        log.warning("Could not import Docling item types; falling back to class-name comparison.")
    return _DOCLING_TYPES


# ── Text normalisation ─────────────────────────────────────────────────────────
def _safe_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (list, tuple)):
        return " ".join(_safe_text(v) for v in value).strip()
    if hasattr(value, "text"):
        return _safe_text(value.text)
    return str(value).strip()


# ── XBRL noise detection ───────────────────────────────────────────────────────
# V1: dual-guard — named patterns (>=2) OR colon-namespace density (>10 in >500 chars)

XBRL_NAMED_PATTERNS = [
    re.compile(r'http[s]?://\S+'),
    re.compile(r'\d{10}\s+\d{4}-\d{2}-\d{2}'),
    re.compile(r'(iso4217|xbrli|us-gaap|srt):[A-Za-z]'),
    re.compile(r'0{7}\d{3}\s+\w+:\w+Member'),
    re.compile(r'<[a-z]+:[A-Za-z]+\s'),
]
_COLON_NS_RE = re.compile(r'\b[a-z][a-z0-9]*:[A-Za-z]')

def is_xbrl_noise(text: str) -> bool:
    if len(text) > 300 and text.count('http') > 4:
        return True
    if sum(1 for p in XBRL_NAMED_PATTERNS if p.search(text)) >= 2:
        return True
    if len(text) > 500 and len(_COLON_NS_RE.findall(text)) > 10:
        return True
    return False


# ── Image filename filter ──────────────────────────────────────────────────────
# W1: filter raw image filename blocks (e.g. aapl-20230202_g1.jpg)

_IMAGE_FILENAME_RE = re.compile(
    r'^\s*[\w\-]+\.(jpg|jpeg|png|gif|svg|webp|tiff|bmp)\s*$', re.I)

def is_image_filename(text: str) -> bool:
    return bool(_IMAGE_FILENAME_RE.match(text))


# ── Boilerplate / footer filter ────────────────────────────────────────────────
# V6 + W3: form footers, page numbers, and filing-layout cover-page fragments.

_BOILERPLATE_RE = re.compile(
    r'(?:'
    r'See\s+accompanying\s+Notes\s+to'
    r'|Apple\s+Inc\.\s*\|\s*Q\d\s*\d{4}'
    r'|\bForm\s+10-[QK]\s*\|\s*\d+\s*$'
    r'|^\s*\|\s*\d+\s*$'
    r')',
    re.I | re.MULTILINE,
)

# W3 + X3: filing-layout cover-page label fragments.
# X3: broadened 'of incorporation)' to also match 'of incorporation or organization)'
_COVER_FRAGMENT_RE = re.compile(
    r'^(?:'
    r'\(State\s+or\s+other\s+jurisdiction'
    r'|of\s+incorporation(?:\s+or\s+organization)?\)'   # X3: added variant
    r'|\(Commission'
    r'|File\s+Number\)'
    r'|\(I\.R\.S\.\s+Employer'
    r'|Identification\s+No\.\)'
    r'|\(Address\s+of\s+principal'
    r'|\(Zip\s+Code\)'
    r'|\(Telephone\s+Number'
    r'|Indicate\s+by\s+check\s+mark\s+whether'
    r')\s*$',
    re.I,
)

# X2: table-spill prose filters — confirmed zero false-positive risk on full 10-Q corpus.
# Parenthesized negative numbers:  ( 1,766 )  ( 12,912 )  ( 0.5 )
_PAREN_NUM_SPILL_RE = re.compile(r'^\s*\(\s*[\d,]+(?:\.\d+)?\s*\)\s*$')
# Bare date column-header fragments:  December 31,   September 24,
_DATE_FRAG_SPILL_RE = re.compile(
    r'^(?:January|February|March|April|May|June|July|August|'
    r'September|October|November|December)\s+\d{1,2},\s*$',
    re.I,
)

def is_boilerplate(text: str) -> bool:
    return bool(_BOILERPLATE_RE.search(text))

def is_cover_fragment(text: str) -> bool:
    """W3/X3: return True for cover-page layout label fragments."""
    return bool(_COVER_FRAGMENT_RE.match(text))

# Y4: column-header word fragments from wide multi-row tables (fair-value, RSU, etc.)
# These isolated single-word or short-phrase "spills" are never meaningful standalone.
# Verified zero false-positive risk on full 10-Q corpus: genuine financial nouns like
# 'Inventories', 'Revenue', 'Net income', 'Goodwill', 'Total assets' are NOT in this list.
_COLUMN_HEADER_SPILL_RE = re.compile(
    r'^(?:'
    r'Adjusted'
    r'|Unrealized'
    r'|Marketable'
    r'|Non-[Cc]urrent'
    r'|Equivalents?'
    r'|Fair\s+Value\s*$'        # standalone "Fair Value" without more context
    r'|Cash\s+and\s*$'          # incomplete phrase ending mid-thought
    r'|Gains?\s*$'              # lone "Gains" or "Gain"
    r'|Losses?\s*$'             # lone "Losses" or "Loss"
    r'|Cost\s*$'                # lone "Cost" (but not "Cost of sales" etc.)
    r')\s*$',
    re.I,
)

def is_table_spill(text: str) -> bool:
    """
    X2 + Y4: return True for unambiguous table-column-overflow prose fragments.
    Patterns:
      X2a: Parenthesized standalone numbers: ( 1,766 )
      X2b: Bare month-day date labels:       December 31,
      Y4:  Known column-header word fragments from wide financial tables:
           Adjusted, Unrealized, Marketable, Non-Current, Equivalents, etc.
    All patterns confirmed zero false-positive risk on full 10-Q prose corpus.
    """
    return bool(
        _PAREN_NUM_SPILL_RE.match(text)
        or _DATE_FRAG_SPILL_RE.match(text)
        or _COLUMN_HEADER_SPILL_RE.match(text)
    )


# ── SEC HTM heading detection ──────────────────────────────────────────────────
# V2: Docling emits SEC HTM section headings as TextItem, not SectionHeaderItem.

_SEC_HEADING_RE = re.compile(
    r'^(?:'
    r'PART\s+[IVX]+\b'
    r'|Item\s+\d+[A-Za-z]?\.'
    r'|Note\s+\d+[\s\-\u2013\u2014]'
    r'|NOTES?\s+TO\s+'
    r'|CONDENSED\s+CONSOLIDATED\s+'
    r')',
    re.I,
)

def _extract_short_heading(text: str) -> Optional[str]:
    if not _SEC_HEADING_RE.match(text.strip()):
        return None
    stripped = text.strip()
    if len(stripped) <= 120:
        # For short blocks, the full text IS the heading.
        # But still trim Note headings that run into body text without a clear break.
        if re.match(r'^Note\s+\d+', stripped, re.I):
            return _extract_note_title(stripped)
        return stripped
    # Long block: extract just the heading prefix.
    if re.match(r'^Note\s+\d+', stripped, re.I):
        return _extract_note_title(stripped)
    m = re.match(
        r'^((?:PART\s+[IVX]+[^.]*?|Item\s+\d+[A-Za-z]?\.\s*[^.]{0,60}|'
        r'CONDENSED\s+CONSOLIDATED\s+[A-Z\s]+))',
        stripped, re.I,
    )
    return m.group(1).strip() if m else stripped[:80]


# Y5: helper — extract clean "Note N - Title" stopping before body text.
_NOTE_BODY_TRIGGER_RE = re.compile(
    r'(?<= )'                               # must be preceded by a space (not mid-word)
    r'\b(The|A\b|An\b|During|In\b|As\b|For\b|This|With|Each|All|These|'
    r'Such|Any|Upon|Under|If\b|When|Pursuant|Effective|Following|Including|'
    r'Certain|At\b|By\b|No\b|Our|Its|From)\b',
    re.I,
)

def _extract_note_title(text: str) -> str:
    """
    Y5: Extract 'Note N - Title' stopping at the first body-text trigger word.
    e.g. 'Note 9 - Contingencies The Company is subject...' -> 'Note 9 - Contingencies'
         'Note 7 - Shareholders Equity Share Repurchase Program During...'
              -> 'Note 7 - Shareholders Equity Share Repurchase Program'
    """
    m = re.match(r'^(Note\s+\d+)\s*[\-\u2013\u2014]\s*', text, re.I)
    if not m:
        return text[:80]
    rest = text[m.end():]
    trigger = _NOTE_BODY_TRIGGER_RE.search(rest)
    title_part = rest[:trigger.start()].strip().rstrip(',').strip() if trigger else rest.strip()
    return (m.group(1) + ' - ' + title_part).rstrip(' -').strip() or text[:80]


# ── Financial-statement title detection ───────────────────────────────────────
# W4: TextItem blocks that are the actual title of a primary financial statement.
# These appear immediately before the statement table and contain the full name.

_FIN_STMT_TITLE_RE = re.compile(
    r'(?:Apple\s+Inc\.\s+)?(?:CONDENSED\s+)?CONSOLIDATED\s+(?:CONDENSED\s+)?'
    r'(?:STATEMENTS?\s+OF\s+(?:OPERATIONS|INCOME|EARNINGS|COMPREHENSIVE\s+INCOME|'
    r'CASH\s+FLOWS?|SHAREHOLDERS|STOCKHOLDERS)|BALANCE\s+SHEET)',
    re.I,
)

def _extract_fin_stmt_title(text: str) -> Optional[str]:
    """
    W4: return a cleaned financial-statement title if text is a statement header,
    e.g. 'Apple Inc. CONDENSED CONSOLIDATED STATEMENTS OF COMPREHENSIVE INCOME'.
    Returns None if text is ordinary prose.
    """
    if not _FIN_STMT_TITLE_RE.search(text):
        return None
    # Keep everything up to the first parenthetical (strips '(Unaudited) (In millions)')
    cleaned = re.sub(r'\s*\((?:Unaudited|In millions[^)]*)\).*', '', text, flags=re.I).strip()
    # Remove 'Apple Inc. ' prefix (redundant given doc_meta)
    cleaned = re.sub(r'^Apple\s+Inc\.\s+', '', cleaned, flags=re.I).strip()
    return cleaned if len(cleaned) >= 10 else None


# ── Cover-admin table detection ────────────────────────────────────────────────
# W10 / W12 / W13: detect cover-page / admin tables that should never be treated
# as financial data tables.

_COVER_ADMIN_RE = re.compile(
    r'(?:'
    # Securities registration table
    r'title\s+of\s+each\s+class'
    r'|trading\s+symbol'
    r'|exchange\s+on\s+which\s+registered'
    # Filer-status checkbox table
    r'|accelerated\s+filer'
    r'|emerging\s+growth\s+company'
    # Table of contents
    r'|^\s*Part\s+[IVX]+\s*/\s*Item\s+\d'
    # Cover identity (address, CIK)
    r'|state\s+or\s+other\s+jurisdiction\s+of\s+incorporation'
    r'|i\.r\.s\.\s+employer\s+identification'
    r')',
    re.I,
)

def is_cover_admin_table(header_text: str, retrieval_text: str) -> bool:
    """W10/W12/W13: return True for cover/admin/TOC tables."""
    combined = f"{header_text} {retrieval_text[:300]}"
    return bool(_COVER_ADMIN_RE.search(combined))


# ── Statement-type classification ─────────────────────────────────────────────
# R4 + W5 + W10/W11: score-based; cover_admin_table short-circuits first.
# W11: equity header_pat tightened to avoid matching securities tables.
# W5: comprehensive income added.

_STATEMENT_SCORING: list[tuple[re.Pattern, re.Pattern, str]] = [
    (
        re.compile(r'consolidated\s+statements?\s+of\s+'
                   r'(operations|income|earnings|profit)', re.I),
        re.compile(r'net\s+(sales|revenue)|gross\s+(profit|margin)|operating\s+income', re.I),
        "income_statement",
    ),
    # W5: comprehensive income — explicit pattern before the generic income pattern
    (
        re.compile(r'consolidated\s+statements?\s+of\s+comprehensive\s+income|'
                   r'other\s+comprehensive\s+income', re.I),
        re.compile(r'comprehensive\s+income|unrealized\s+(gains?|losses?)', re.I),
        "comprehensive_income_statement",
    ),
    (
        re.compile(r'consolidated\s+balance\s+sheet|'
                   r'consolidated\s+statements?\s+of\s+financial\s+position', re.I),
        re.compile(r'total\s+assets|total\s+liabilities', re.I),
        "balance_sheet",
    ),
    (
        re.compile(r'consolidated\s+statements?\s+of\s+cash\s+flows?', re.I),
        re.compile(r'cash\s+(used|provided|from)\s+(in|by)\s+'
                   r'(operating|investing|financing)', re.I),
        "cash_flow_statement",
    ),
    (
        re.compile(r'consolidated\s+statements?\s+of\s+(stockholders|shareholders).equity|'
                   r'changes\s+in\s+(stockholders|shareholders).equity', re.I),
        # W11: tightened — "common stock AND additional" to avoid matching
        # securities registration table which also says "Common Stock"
        re.compile(r'common\s+stock\s+and\s+additional|retained\s+earnings|'
                   r'accumulated\s+other\s+comprehensive', re.I),
        "equity_statement",
    ),
    (
        re.compile(r'segment\s+(information|results|data|revenue)|'
                   r'(revenue|income)\s+by\s+(segment|geography|region)', re.I),
        re.compile(r'segment|geographic|region', re.I),
        "segment_table",
    ),
    (
        re.compile(r'earnings\s+per\s+share|per\s+share\s+data', re.I),
        re.compile(r'\beps\b|basic|diluted', re.I),
        "eps_table",
    ),
    (
        re.compile(r'(long.term\s+)?debt|borrowings?|'
                   r'notes?\s+payable|credit\s+facilit', re.I),
        re.compile(r'maturity|interest\s+rate|principal|coupon', re.I),
        "debt_table",
    ),
    (
        re.compile(r'stock.based\s+compensation|share.based|'
                   r'stock\s+options?|restricted\s+stock\s+units?', re.I),
        re.compile(r'grant|vest|exercise|fair\s+value', re.I),
        "compensation_table",
    ),
]

def classify_statement_type(
    title: str, header_text: str, retrieval_text: str = ""
) -> str:
    """
    Score-based classification.  cover_admin_table short-circuits the full scoring.
    Returns 'cover_admin_table' for registration, checkbox, and TOC tables.
    """
    # W10/W12/W13: cover-admin takes priority
    if is_cover_admin_table(header_text, retrieval_text):
        return "cover_admin_table"

    best_label = "other_table"
    best_score = 0
    for title_pat, header_pat, label in _STATEMENT_SCORING:
        score = (3 if title_pat.search(title) else 0) + (1 if header_pat.search(header_text) else 0)
        if score > best_score:
            best_score, best_label = score, label
    return best_label


# ── Period-type detection ──────────────────────────────────────────────────────
# R6 + W9: added maturity_schedule detection.

_PERIOD_ANNUAL_RE     = re.compile(
    r'year\s+ended|annual|twelve\s+months|52[- ]weeks?|53[- ]weeks?', re.I)
_PERIOD_QTRLY_RE      = re.compile(
    r'quarter(?:ly)?|three\s+months|thirteen\s+weeks|nine\s+months|six\s+months', re.I)
_PERIOD_INSTANT_RE    = re.compile(
    # X5: requires 'as of' or 'balance at' context so bare 'Date: February 2, 2023'
    # headers no longer trigger an 'instant' period signal.
    r'as\s+of\b|balance\s+at\b',
    re.I)
# W9: maturity-bucket schedules — year buckets like "due after 1 year", "2025", "Thereafter"
_PERIOD_MATURITY_RE   = re.compile(
    r'due\s+(after|within|in)|thereafter|remaining\s+(months?|year)|'
    r'maturities|maturity\s+date', re.I)
# Y1: complete single-line filing dates (Month D, YYYY with no embedded newline).
# These appear as table headers in signature blocks and repurchase tables and must
# not be treated as reporting-period column headers.  Genuine period column headers
# always have the year on a second line: 'December 31,\n2022' (newline, not space).
_FILING_DATE_RE = re.compile(
    r'^(?:January|February|March|April|May|June|July|August|'
    r'September|October|November|December)'
    r'\s+\d{1,2},[ \t]+\d{4}$',   # space/tab between comma and year — NOT a newline
    re.I,
)

def detect_period_info(header_row: list[str], row_labels: list[str] | None = None) -> dict:
    """
    Return {period_type, period_signals}.
    W9: maturity_schedule takes priority when maturity signals are present in
    headers OR row_labels (year-bucket strings like 2024, 2025, Thereafter).
    """
    joined = " ".join(header_row)
    # W9: check row labels too for maturity buckets
    row_joined = " ".join(row_labels or [])

    signals = []
    if _PERIOD_MATURITY_RE.search(joined) or _PERIOD_MATURITY_RE.search(row_joined):
        signals.append("maturity_schedule")
    if _PERIOD_ANNUAL_RE.search(joined):   signals.append("annual")
    if _PERIOD_QTRLY_RE.search(joined):    signals.append("quarterly")
    if _PERIOD_INSTANT_RE.search(joined):  signals.append("instant")

    # maturity_schedule takes priority over quarterly when both fire
    if "maturity_schedule" in signals:
        period_type = "maturity_schedule"
    elif len(signals) == 0:
        period_type = "unknown"
    elif len(signals) == 1:
        period_type = signals[0]
    else:
        period_type = "mixed"

    return {"period_type": period_type, "period_signals": signals}


# ── Financial-statement and cover-page flags ───────────────────────────────────
# W6: _FIN_INTEGER_RE fixed to match comma-formatted numbers (72,994 etc.)

# Matches: $1,234  or  $12345  or  1,234,567
_FIN_DOLLAR_RE   = re.compile(r'\$\s*[\d,]{4,}')
# W6 fix: match numbers with optional thousands-commas, 4+ total digits
_FIN_INTEGER_RE  = re.compile(r'\b\d{1,3}(?:,\d{3})+\b|\b\d{5,}\b')
_COVER_TEXT_RE   = re.compile(r'\$0\.\d{5}|par\s+value|CUSIP|ISIN', re.I)

def is_financial_statement_table(statement_type: str, grid: list[list[str]]) -> bool:
    if statement_type in ("income_statement", "balance_sheet", "cash_flow_statement",
                          "equity_statement", "comprehensive_income_statement"):
        return True
    if statement_type == "cover_admin_table":
        return False
    all_text = " ".join(cell for row in grid for cell in row)
    if _COVER_TEXT_RE.search(all_text):
        return False
    return (bool(_FIN_DOLLAR_RE.search(all_text)) or
            bool(_FIN_INTEGER_RE.search(all_text))) and len(grid) >= 3


def is_cover_page_table(grid: list[list[str]]) -> bool:
    """Soft heuristic; only fires on tiny tables with no financial numbers."""
    all_text    = " ".join(cell for row in grid for cell in row)
    total_cells = sum(len(row) for row in grid)
    return not _FIN_INTEGER_RE.search(all_text) and total_cells <= 12


# ── Converter singleton ────────────────────────────────────────────────────────
_CONVERTER      = None
_CONVERTER_LOCK = threading.Lock()

def _build_converter():
    from docling.document_converter import DocumentConverter
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.datamodel.base_models import InputFormat
    from docling.document_converter import PdfFormatOption, WordFormatOption

    return DocumentConverter(
        format_options={
            InputFormat.PDF:  PdfFormatOption(
                pipeline_options=PdfPipelineOptions(do_ocr=True, do_table_structure=True)
            ),
            InputFormat.DOCX: WordFormatOption(),
        }
    )

def get_converter():
    global _CONVERTER
    if _CONVERTER is None:
        with _CONVERTER_LOCK:
            if _CONVERTER is None:
                log.info("Initialising Docling converter (one-time load)...")
                _CONVERTER = _build_converter()
    return _CONVERTER


# ── Page provenance ────────────────────────────────────────────────────────────
def get_page_range(item) -> tuple[int, int, list[int]]:
    try:
        if hasattr(item, "prov") and item.prov:
            pages = sorted({p.page_no for p in item.prov})
            return pages[0], pages[-1], pages
    except Exception:
        pass
    return 1, 1, [1]


# ── Table grid extraction ──────────────────────────────────────────────────────
def _extract_grid(table) -> list[list[str]]:
    try:
        td       = table.data
        cells    = td.table_cells
        num_rows = td.num_rows
        num_cols = td.num_cols
        if not cells or num_rows == 0 or num_cols == 0:
            return []

        grid   = [[""] * num_cols for _ in range(num_rows)]
        filled = [[False]  * num_cols for _ in range(num_rows)]

        for cell in cells:
            r, c = cell.start_row_offset_idx, cell.start_col_offset_idx
            if r >= num_rows or c >= num_cols or filled[r][c]:
                continue
            text = _safe_text(cell.text)
            if text in ("<!-- rich cell -->", "<!--rich cell-->"):
                text = ""
            grid[r][c]   = text
            filled[r][c] = True

        non_empty_cols = [c for c in range(num_cols)
                          if any(grid[r][c] for r in range(num_rows))]
        if not non_empty_cols:
            return []
        grid = [[row[c] for c in non_empty_cols] for row in grid]

        seen, deduped = set(), []
        for row in grid:
            key = "|".join(row)
            if key not in seen and any(cell for cell in row):
                seen.add(key)
                deduped.append(row)
        return deduped

    except Exception as e:
        log.warning(f"    _extract_grid failed: {e}")
        return []


# ── Header / data boundary ─────────────────────────────────────────────────────
_NUM_CELL_RE  = re.compile(r'^\s*[\$\(]?[\d,]+\.?\d*[BKMG%\)]?\s*$')
_PURE_NUM_RE  = re.compile(r'^\s*[\(\-]?[\d,]+\.?\d*\)?\s*$')
_UNIT_TOKEN_RE = re.compile(r'^\s*[\$%€£¥]\s*$')


def _row_numeric_density(row: list[str]) -> float:
    non_empty = [c for c in row if c]
    if not non_empty:
        return 0.0
    return sum(1 for c in non_empty if _NUM_CELL_RE.match(c)) / len(non_empty)


def _split_header_and_data(
    grid: list[list[str]],
) -> tuple[list[str], list[list[str]], list[str]]:
    """
    Returns (merged_header, data_rows, units).
    V4: ejects absorbed numeric cells to synthetic first data row.
    V5: strips standalone unit tokens into 'units' list.
    """
    if not grid:
        return [], [], []

    header_row_count = 1
    for i in range(1, min(4, len(grid))):
        if _row_numeric_density(grid[i]) < 0.5:
            header_row_count += 1
        else:
            break

    header_rows = grid[:header_row_count]
    data_rows   = list(grid[header_row_count:])
    num_cols    = len(header_rows[0])
    merged_header = []
    units         = []
    absorbed_row  = [""] * num_cols

    for c in range(num_cols):
        parts, col_units, col_absorbed = [], [], ""
        for r in range(header_row_count):
            val = header_rows[r][c] if c < len(header_rows[r]) else ""
            if not val:
                continue
            if _PURE_NUM_RE.match(val):
                col_absorbed = val
            elif _UNIT_TOKEN_RE.match(val):
                col_units.append(val.strip())
            else:
                parts.append(val)
        merged_header.append(" / ".join(parts) if parts else "")
        units.extend(col_units)
        absorbed_row[c] = col_absorbed

    if any(absorbed_row):
        data_rows.insert(0, absorbed_row)

    return merged_header, data_rows, list(dict.fromkeys(units))


# ── Phantom column removal ─────────────────────────────────────────────────────
# W7: drop columns where header == '' AND every data cell is also ''.
# These are artefacts left by V5's unit-stripping.

def _drop_phantom_columns(
    header: list[str],
    data_rows: list[list[str]],
) -> tuple[list[str], list[list[str]]]:
    """Remove columns that are entirely empty in both header and data."""
    if not header:
        return header, data_rows
    keep = []
    for c in range(len(header)):
        if header[c]:
            keep.append(c)
        elif any(row[c] for row in data_rows if c < len(row)):
            keep.append(c)  # empty header but has data → keep
        # else: empty header, all-empty data → phantom, drop
    if len(keep) == len(header):
        return header, data_rows  # nothing to drop
    new_header    = [header[c] for c in keep]
    new_data_rows = [[row[c] for c in keep if c < len(row)] for row in data_rows]
    return new_header, new_data_rows


def _compact_value_columns(
    header: list[str],
    data_rows: list[list[str]],
) -> tuple[list[str], list[list[str]]]:
    """
    Y2: Fold empty-header columns that carry data into their preceding non-empty column.

    Root cause: SEC financial statement HTM tables have a structure like:
        [label | period_hdr | $ | value | period_hdr | $ | value]
    where the '$' row is part of the data section (not a header row).
    After V5 strips '$' from header positions, the pattern becomes:
        [label | period_hdr | '' | period_hdr | '']
    with the actual numeric data in the '' columns.

    _drop_phantom_columns correctly keeps these '' columns (they have data).
    But cell_grid then gets col_header='' for the real financial values.

    This function detects the pattern [non-empty-header, empty-header] and remaps
    the empty-header column's data to the preceding non-empty column, then drops
    the empty-header column. Result: each period column header directly carries its values.
    """
    if not header:
        return header, data_rows

    n = len(header)
    # Build owner map: empty-header cols map to their preceding non-empty col
    col_owner = list(range(n))
    for c in range(1, n):
        if not header[c] and header[c - 1]:
            col_owner[c] = c - 1

    # If nothing to remap, return as-is
    if col_owner == list(range(n)):
        return header, data_rows

    # Determine which columns to keep (only "owner" columns)
    keep_set = set(col_owner)
    keep_cols = sorted(keep_set)

    new_header = [header[c] for c in keep_cols]
    keep_idx = {c: i for i, c in enumerate(keep_cols)}

    new_data: list[list[str]] = []
    for row in data_rows:
        new_row = [''] * len(keep_cols)
        for orig_c in range(n):
            owner_c = col_owner[orig_c]
            new_c   = keep_idx[owner_c]
            val = row[orig_c] if orig_c < len(row) else ''
            if val and not _UNIT_TOKEN_RE.match(val):
                # Skip standalone unit tokens ($, %) so they don't block the
                # real numeric value that follows in the adjacent '' column.
                if not new_row[new_c]:
                    new_row[new_c] = val
        new_data.append(new_row)

    return new_header, new_data


# ── Row-label column inference ─────────────────────────────────────────────────
_YEAR_HEADER_RE = re.compile(r'\b(19|20)\d{2}\b|Q[1-4]|quarter|ended', re.I)

def _infer_label_col(header: list[str], data_rows: list[list[str]]) -> int:
    for col_idx in range(len(header)):
        if _YEAR_HEADER_RE.search(header[col_idx]):
            continue
        values = [row[col_idx] for row in data_rows
                  if col_idx < len(row) and row[col_idx]]
        if not values:
            continue
        if sum(1 for v in values if not _NUM_CELL_RE.match(v)) / len(values) >= 0.6:
            return col_idx
    return 0


# ── Table-text renderer ────────────────────────────────────────────────────────
# X6: accepts label_col so the label column's merged header is never used as a
# row-value prefix. Previously 'ASSETS: / Current assets: / Cash and cash equivalents'
# was prepended to every balance sheet data row even when the row belonged to
# the liabilities or equity section.

def _render_table_text(
    header: list[str],
    data_rows: list[list[str]],
    label_col: int = 0,
) -> str:
    lines = []
    header_line = " | ".join(h for h in header if h)
    if header_line:
        lines.append(header_line)
    for row in data_rows:
        parts = []
        for col_idx, val in enumerate(row):
            if not val:
                continue
            # X6: skip the label column's merged header as a prefix for other cells.
            # The label column value itself is used as the row identifier and should
            # appear as a plain value, not as "ASSETS: / ... : Services".
            if col_idx == label_col:
                parts.append(val)
                continue
            col_name = header[col_idx] if col_idx < len(header) else ""
            parts.append(f"{col_name}: {val}" if col_name and col_name != val else val)
        if parts:
            lines.append(" | ".join(parts))
    return "\n".join(lines)


# ── Full table record builder ──────────────────────────────────────────────────

def build_table_record(
    table,
    page_start:        int,
    page_end:          int,
    pages:             list[int],
    page_is_surrogate: bool,
    section:           str,
    table_title:       str,
    doc_meta:          dict,
) -> Optional[dict]:
    try:
        grid = _extract_grid(table)
        if not grid:
            return None

        header, data_rows, units = _split_header_and_data(grid)
        if not header:
            return None

        # W7: remove phantom empty columns left by unit stripping
        header, data_rows = _drop_phantom_columns(header, data_rows)
        if not header:
            return None

        # Y2: fold empty-header data columns into their preceding period column.
        # This fixes cell_grid entries with col_header='' that occur because the
        # $ companion column carries the actual value while the period column carries '$'.
        header, data_rows = _compact_value_columns(header, data_rows)
        if not header:
            return None

        header_text    = " ".join(h for h in header if h)
        # Infer label column before rendering so the renderer can skip its header as prefix (X6)
        label_col  = _infer_label_col(header, data_rows)
        row_labels = [row[label_col] if label_col < len(row) else "" for row in data_rows]

        # X6: pass label_col so the label column header is not used as a row-value prefix
        retrieval_text = _render_table_text(header, data_rows, label_col)
        if not retrieval_text:
            return None

        statement_type = classify_statement_type(table_title, header_text, retrieval_text)

        # W9: pass row_labels to detect maturity-bucket schedules
        period_info = detect_period_info(header, row_labels)

        # W8: filter standalone unit tokens from data cell values
        cell_grid = []
        for row in data_rows:
            row_label = row[label_col] if label_col < len(row) else ""
            for col_idx, val in enumerate(row):
                if col_idx == label_col or not val:
                    continue
                if _UNIT_TOKEN_RE.match(val):   # W8: skip bare $, % etc. in data
                    continue
                if col_idx < len(header):
                    cell_grid.append({
                        "row_label":  row_label,
                        "col_header": header[col_idx],
                        "value":      val,
                    })

        # X4: cover_admin_table tables (registration, TOC, checkboxes) must have
        # periods=[] — year-like strings in debt maturity names such as "Notes due 2024"
        # must not be misread as reporting-period columns.
        if statement_type == "cover_admin_table":
            period_cols = []
        else:
            # Y1: exclude headers that are complete single-line filing dates
            # (e.g. 'February 2, 2023') from period_cols. Genuine period column
            # headers always have the year on a separate line ('December 31,\n2022')
            # or are prefixed with a period phrase ('Three Months Ended / ...').
            period_cols = [
                h for h in header
                if (re.search(r'\b(19|20)\d{2}\b', h)
                    or re.search(r'Q[1-4]|quarter|months|year|ended|fiscal', h, re.I))
                and not _FILING_DATE_RE.match(h)   # Y1: not a bare filing date
            ]

        # Y3: cover_admin_table tables are definitionally cover-page content —
        # force is_cover_page=True for consistency regardless of numeric content.
        if statement_type == "cover_admin_table":
            is_cover = True
        else:
            is_cover = is_cover_page_table(grid)
        is_fin_stm = is_financial_statement_table(statement_type, grid)

        return {
            "retrieval_text":          retrieval_text,
            "column_headers":          header,
            "row_labels":              row_labels,
            "label_col_index":         label_col,
            "cell_grid":               cell_grid,
            "periods":                 period_cols,
            "units":                   units,
            "table_title":             table_title,
            "statement_type":          statement_type,
            "period_type":             period_info["period_type"],
            "period_signals":          period_info["period_signals"],
            "is_financial_statement":  is_fin_stm,
            "is_cover_page":           is_cover,
            "page":                    page_start,
            "page_start":              page_start,
            "page_end":                page_end,
            "pages":                   pages,
            "page_is_surrogate":       page_is_surrogate,
            "section":                 section,
            **doc_meta,
        }

    except Exception as e:
        log.warning(f"    build_table_record failed: {e}")
        return None


# ── Figure / image handling ────────────────────────────────────────────────────
_FIGURE_FALLBACK = "[Figure present -- no caption or metadata available]"

def _describe_figure(item, caption: str) -> str:
    parts = [caption] if caption else []
    try:
        if hasattr(item, "annotations") and item.annotations:
            for ann in item.annotations:
                ann_text = _safe_text(getattr(ann, "text", None) or ann)
                if ann_text and ann_text not in parts:
                    parts.append(ann_text)
    except Exception:
        pass
    try:
        if hasattr(item, "image") and item.image:
            img = item.image
            if hasattr(img, "uri") and img.uri:
                parts.append(f"[Figure URI: {img.uri}]")
            elif hasattr(img, "pil_image") and img.pil_image is not None:
                w, h = img.pil_image.size
                parts.append(f"[Figure: {w}x{h}px]")
    except Exception:
        pass
    if not parts:
        return _FIGURE_FALLBACK
    return " | ".join(parts)


# ── Excel handler ──────────────────────────────────────────────────────────────
def parse_excel(raw_path: Path, doc_meta: dict) -> Optional[dict]:
    if raw_path.suffix.lower() == ".xls":
        log.error(f"  Legacy .xls not supported. Convert {raw_path.name} to .xlsx first.")
        return None
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl")
        return None
    try:
        wb = openpyxl.load_workbook(str(raw_path), data_only=True)
    except Exception as e:
        log.error(f"  openpyxl failed: {e}")
        return None

    table_records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state != "visible":
            continue
        raw_rows = []
        for row in ws.iter_rows(values_only=True):
            str_row = [str(c) if c is not None else "" for c in row]
            if any(c for c in str_row):
                raw_rows.append(str_row)
        if not raw_rows:
            continue

        header, data_rows, units = _split_header_and_data(raw_rows)
        if not header or not data_rows:
            continue
        header, data_rows = _drop_phantom_columns(header, data_rows)
        header, data_rows = _compact_value_columns(header, data_rows)   # Y2

        label_col      = _infer_label_col(header, data_rows)
        retrieval_text = _render_table_text(header, data_rows, label_col)
        row_labels     = [row[label_col] if label_col < len(row) else "" for row in data_rows]
        period_info    = detect_period_info(header, row_labels)
        statement_type = classify_statement_type(sheet_name, " ".join(header), retrieval_text)

        cell_grid = []
        for row in data_rows:
            row_label = row[label_col] if label_col < len(row) else ""
            for col_idx, val in enumerate(row):
                if col_idx == label_col or not val or _UNIT_TOKEN_RE.match(val):
                    continue
                if col_idx < len(header):
                    cell_grid.append({"row_label": row_label,
                                      "col_header": header[col_idx], "value": val})

        # X4 + Y1: cover_admin_table gets empty periods; filing dates excluded from period_cols
        if statement_type == "cover_admin_table":
            period_cols = []
            is_cover    = True    # Y3: cover_admin is always cover-page content
        else:
            period_cols = [
                h for h in header
                if (re.search(r'\b(19|20)\d{2}\b', h)
                    or re.search(r'Q[1-4]|quarter|months|year|ended|fiscal', h, re.I))
                and not _FILING_DATE_RE.match(h)
            ]
            is_cover = is_cover_page_table([[c for c in row] for row in (data_rows or [[]])])

        table_records.append({
            "retrieval_text":         retrieval_text,
            "column_headers":         header,
            "row_labels":             row_labels,
            "label_col_index":        label_col,
            "cell_grid":              cell_grid,
            "periods":                period_cols,
            "units":                  units,
            "table_title":            sheet_name,
            "statement_type":         statement_type,
            "period_type":            period_info["period_type"],
            "period_signals":         period_info["period_signals"],
            "is_financial_statement": bool(_FIN_INTEGER_RE.search(retrieval_text)),
            "is_cover_page":          is_cover,
            "page": 1, "page_start": 1, "page_end": 1, "pages": [1],
            "page_is_surrogate":      False,
            "section":                sheet_name,
            **doc_meta,
        })

    if not table_records:
        log.warning(f"  No usable sheets in {raw_path.name}")
        return None

    output = {**doc_meta, "raw_path": str(raw_path),
               "prose_blocks": [], "table_records": table_records,
               "stats": {"prose_count": 0, "table_count": len(table_records)}}
    log.info(f"  Excel parsed: {len(table_records)} sheet(s)")
    return output


# ── Core parse function ────────────────────────────────────────────────────────

def parse_document(raw_path: str, doc_meta: dict) -> Optional[dict]:
    p = Path(raw_path)
    if not p.is_absolute():
        p = BASE_DIR / p
    raw_path = p

    if not raw_path.exists():
        log.error(f"  File not found: {raw_path}")
        return None

    suffix = raw_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return parse_excel(raw_path, doc_meta)

    log.info(f"  Parsing {raw_path.name} ...")
    try:
        result = get_converter().convert(str(raw_path))
        doc    = result.document
    except Exception as e:
        log.error(f"  Docling conversion failed for {raw_path.name}: {e}")
        return None

    is_htm = suffix in (".htm", ".html")
    dtypes = _get_docling_types()

    def _is_type(item, name: str) -> bool:
        cls = dtypes.get(name)
        return isinstance(item, cls) if cls is not None else type(item).__name__ == name

    prose_blocks        = []
    table_records       = []
    current_section     = "General"
    pending_table_title = "General"
    block_order         = 0

    try:
        for item, level in doc.iterate_items():
            block_order += 1

            # ── Section header (native Docling type) ──────────────────────────
            if _is_type(item, "SectionHeaderItem"):
                text = _safe_text(item.text)
                if text:
                    current_section     = text
                    pending_table_title = text

            # ── Text block ────────────────────────────────────────────────────
            elif _is_type(item, "TextItem"):
                text = _safe_text(item.text)
                if len(text) < 8:
                    continue
                if is_xbrl_noise(text):       # V1
                    continue
                if is_boilerplate(text):       # V6
                    continue
                if is_image_filename(text):    # W1: filter raw image filenames
                    continue
                if is_cover_fragment(text):    # W3/X3: filter cover layout fragments
                    continue
                if is_table_spill(text):       # X2: filter paren-negatives and date fragments
                    continue

                # W4 + X1: detect financial-statement title TextItems.
                # X1: also resets current_section to correct the 'Item 1A.' TOC bleed.
                # The TOC emits SectionHeaderItems ('Item 1A.', etc.) that persist until
                # the first actual financial statement header appears. When we detect a
                # primary-statement title, we know we are now in Item 1 Financial Statements.
                fin_title = _extract_fin_stmt_title(text)
                if fin_title:
                    pending_table_title = fin_title
                    current_section     = "Item 1 - Financial Statements"  # X1
                    # Don't store as prose (it's a table header, not body text)
                    continue

                # V2: detect SEC HTM headings emitted as TextItem
                heading = _extract_short_heading(text)
                if heading:
                    current_section     = heading
                    pending_table_title = heading
                    if len(text.strip()) <= 120:
                        continue
                    body = text[len(heading):].strip()
                    if len(body) < 8:
                        continue
                    text = body

                ps, pe, pages = get_page_range(item)
                prose_blocks.append({
                    "text":              text,
                    "page":              ps,
                    "page_start":        ps,
                    "page_end":          pe,
                    "pages":             pages,
                    "page_is_surrogate": is_htm,
                    "section":           current_section,
                    "type":              "prose",
                    "block_order":       block_order,
                })

            # ── Table ─────────────────────────────────────────────────────────
            elif _is_type(item, "TableItem"):
                ps, pe, pages = get_page_range(item)
                title = pending_table_title
                try:
                    cap = _safe_text(item.caption)
                    if cap:
                        title = cap
                except Exception:
                    pass

                record = build_table_record(
                    item,
                    page_start=ps, page_end=pe, pages=pages,
                    page_is_surrogate=is_htm,
                    section=current_section,
                    table_title=title,
                    doc_meta=doc_meta,
                )
                if record:
                    record["block_order"] = block_order
                    table_records.append(record)
                    pending_table_title = current_section  # reset after use

            # ── Figure / picture ─────────────────────────────────────────────
            elif _is_type(item, "PictureItem"):
                caption = ""
                try:
                    caption = _safe_text(item.caption) or _safe_text(item.text)
                except Exception:
                    pass

                description = _describe_figure(item, caption)

                # W2: drop figures that have nothing but the fallback placeholder
                if description == _FIGURE_FALLBACK:
                    continue

                ps, pe, pages = get_page_range(item)
                prose_blocks.append({
                    "text":              f"[Figure] {description}",
                    "page":              ps,
                    "page_start":        ps,
                    "page_end":          pe,
                    "pages":             pages,
                    "page_is_surrogate": is_htm,
                    "section":           current_section,
                    "type":              "figure_caption",
                    "block_order":       block_order,
                })

    except Exception as e:
        log.warning(f"  Item iteration partial failure: {e}")

    if not prose_blocks and not table_records:
        log.warning(f"  No content extracted from {raw_path.name}")
        return None

    # #17: content sanity check
    try:
        file_size_kb = raw_path.stat().st_size / 1024
        min_expected = max(5, int(file_size_kb * 0.02))
        if len(prose_blocks) < min_expected and suffix == ".pdf":
            log.warning(
                f"  Sanity check: only {len(prose_blocks)} prose blocks from "
                f"{file_size_kb:.0f} KB PDF (expected >={min_expected})."
            )
    except Exception:
        pass

    return {
        "doc_id":        doc_meta["doc_id"],
        "ticker":        doc_meta["ticker"],
        "company":       doc_meta["company"],
        "form_type":     doc_meta["form_type"],
        "filing_date":   doc_meta["filing_date"],
        "fiscal_year":   doc_meta["fiscal_year"],
        "source_url":    doc_meta["source_url"],
        "raw_path":      str(raw_path),
        "prose_blocks":  prose_blocks,
        "table_records": table_records,
        "stats": {
            "prose_count": len(prose_blocks),
            "table_count": len(table_records),
        },
    }


# ── Atomic JSON write ──────────────────────────────────────────────────────────
def _atomic_json_write(path: Path, data) -> None:
    tmp = path.with_suffix(".tmp")
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        tmp.replace(path)
    except Exception as e:
        log.error(f"  Atomic write failed for {path}: {e}")
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise

def save_parsed(output: dict) -> Path:
    doc_id      = output["doc_id"]
    parsed_path = PARSED_DIR / f"{doc_id}.json"
    _atomic_json_write(parsed_path, output)
    if output["table_records"]:
        _atomic_json_write(TABLES_DIR / f"{doc_id}_tables.json", output["table_records"])
    return parsed_path


# ── Manifest helpers ───────────────────────────────────────────────────────────
def load_manifest() -> list[dict]:
    if not MANIFEST.exists():
        return []
    entries, bad = [], 0
    with open(MANIFEST, encoding="utf-8") as f:
        for line_no, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError as e:
                log.warning(f"  Manifest parse error at line {line_no}: {e}")
                bad += 1
    if bad:
        log.warning(f"  {bad} malformed line(s) skipped in manifest.")
    return entries

def _write_manifest(entries: list[dict]) -> None:
    tmp = MANIFEST.with_suffix(".tmp")
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            for e in entries:
                f.write(json.dumps(e) + "\n")
        tmp.replace(MANIFEST)
    except Exception as e:
        log.error(f"  Manifest write failed: {e}")
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise


# ── Entry point ────────────────────────────────────────────────────────────────
def run(tickers: list[str] | None = None, reparse: bool = False) -> None:
    """O(N) manifest handling: load once, update in-memory, write once."""
    with _MANIFEST_LOCK:
        all_entries = load_manifest()

    id_to_idx = {e.get("doc_id"): i for i, e in enumerate(all_entries)}
    working   = all_entries
    if tickers:
        tickers_upper = {t.strip().upper() for t in tickers}
        working = [e for e in all_entries if e.get("ticker", "").upper() in tickers_upper]

    pending = [e for e in working if reparse or e.get("parse_status") == "pending"]
    log.info(f"Parsing {len(pending)} documents (reparse={reparse})")

    success = failed = 0
    for entry in pending:
        doc_id = entry["doc_id"]
        rp     = Path(entry["raw_path"])
        if not rp.is_absolute():
            rp = BASE_DIR / rp

        parsed_path = PARSED_DIR / f"{doc_id}.json"
        if parsed_path.exists() and not reparse:
            log.info(f"  Already parsed: {doc_id}")
            idx = id_to_idx.get(doc_id)
            if idx is not None:
                all_entries[idx]["parse_status"] = "parsed"
            success += 1
            continue

        doc_meta = {
            "doc_id":      doc_id,
            "ticker":      entry["ticker"],
            "company":     entry.get("company", entry["ticker"]),
            "form_type":   entry["form_type"],
            "filing_date": entry["filing_date"],
            "fiscal_year": entry["fiscal_year"],
            "source_url":  entry.get("source_url", ""),
        }

        output     = parse_document(str(rp), doc_meta)
        new_status = "parsed" if output else "failed"
        idx        = id_to_idx.get(doc_id)
        if idx is not None:
            all_entries[idx]["parse_status"] = new_status

        if output:
            save_parsed(output)
            success += 1
        else:
            failed += 1

    with _MANIFEST_LOCK:
        _write_manifest(all_entries)

    log.info(f"\nParsing complete. Success: {success} | Failed: {failed}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Parse SEC filings with Docling.")
    parser.add_argument("--tickers", nargs="+")
    parser.add_argument("--reparse", action="store_true")
    args = parser.parse_args()
    run(tickers=args.tickers, reparse=args.reparse)